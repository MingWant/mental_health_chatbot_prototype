"""
Mental Health RAG Service
Specialized in handling mental health related documents and knowledge base
"""

import os
import uuid
import asyncio
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import chromadb
from chromadb.config import Settings
import aiofiles
import chardet
import jieba
from datetime import datetime
import json

# Document processing related
import PyPDF2
from docx import Document
import openpyxl
from sentence_transformers import SentenceTransformer

class MentalHealthDocumentProcessor:
    """Mental Health Document Processor"""
    
    def __init__(self):
        self.supported_extensions = {'.txt', '.pdf', '.docx', '.xlsx', '.md'}
        
        # Mental health related keywords and categories
        self.mental_health_categories = {
            # Core Emotional Health
            "Emotion Management": ["emotion", "mood", "anxiety", "depression", "anger", "stress", "happiness", "sadness", "joy", "fear", "worry", "panic", "grief", "loneliness", "hopelessness", "guilt", "shame", "frustration", "irritability", "emotional regulation", "emotional intelligence"],
            
            # Stress & Pressure Management
            "Stress Management": ["stress", "tension", "fatigue", "burnout", "pressure", "overwhelm", "exhaustion", "chronic stress", "acute stress", "stress response", "fight or flight", "cortisol", "adrenaline", "stress hormones", "work stress", "academic stress", "financial stress"],
            
            # Sleep & Rest
            "Sleep Health": ["sleep", "insomnia", "dreams", "rest", "sleep quality", "sleep hygiene", "sleep disorders", "sleep apnea", "restless legs", "nightmares", "sleep paralysis", "circadian rhythm", "sleep schedule", "sleep deprivation", "napping", "bedtime routine", "sleep environment"],
            
            # Academic & Learning Mental Health
            "Learning Mental Health": ["learning", "exams", "grades", "academics", "focus", "memory", "study skills", "test anxiety", "academic pressure", "performance anxiety", "learning disabilities", "ADHD", "dyslexia", "concentration", "attention", "cognitive load", "mental fatigue", "brain fog"],
            
            # Social & Relationship Health
            "Interpersonal Relationships": ["friends", "family", "social", "communication", "relationships", "loneliness", "social anxiety", "peer pressure", "bullying", "conflict resolution", "boundaries", "trust", "intimacy", "attachment", "social skills", "social isolation", "social support", "community"],
            
            # Self-Care & Wellness
            "Self-Care": ["self-care", "self-compassion", "relaxation", "meditation", "mindfulness", "self-love", "self-acceptance", "self-worth", "self-esteem", "self-image", "body image", "personal growth", "self-improvement", "wellness", "balance", "me time", "self-soothing"],
            
            # Therapeutic Approaches
            "Cognitive Behavioral Therapy": ["CBT", "cognition", "behavior", "thinking", "cognitive distortion", "automatic thoughts", "cognitive restructuring", "behavioral activation", "exposure therapy", "thought records", "cognitive therapy", "rational emotive therapy", "cognitive patterns", "negative thinking"],
            
            "Mindfulness Meditation": ["mindfulness", "meditation", "breathing", "awareness", "present moment", "mindful living", "mindful eating", "mindful walking", "body scan", "loving-kindness", "compassion meditation", "vipassana", "zen", "transcendental meditation", "guided meditation"],
            
            # Crisis & Emergency
            "Crisis Intervention": ["suicide", "self-harm", "crisis", "emergency", "help", "crisis intervention", "suicidal thoughts", "suicidal ideation", "self-injury", "cutting", "overdose", "crisis hotline", "emergency services", "safety planning", "risk assessment", "protective factors"],
            
            # Trauma & PTSD
            "Trauma & PTSD": ["trauma", "PTSD", "post-traumatic stress", "traumatic experience", "flashbacks", "nightmares", "hypervigilance", "avoidance", "emotional numbness", "dissociation", "complex trauma", "childhood trauma", "sexual trauma", "physical trauma", "emotional trauma", "trauma therapy", "EMDR"],
            
            # Anxiety Disorders
            "Anxiety Disorders": ["generalized anxiety", "panic disorder", "social anxiety", "phobias", "agoraphobia", "separation anxiety", "health anxiety", "hypochondria", "obsessive-compulsive", "OCD", "intrusive thoughts", "compulsions", "anxiety attacks", "panic attacks", "anxiety symptoms", "anxiety treatment"],
            
            # Mood Disorders
            "Mood Disorders": ["depression", "major depression", "bipolar disorder", "manic depression", "cyclothymia", "dysthymia", "seasonal depression", "postpartum depression", "mood swings", "mania", "hypomania", "depressive episodes", "mood stabilizers", "antidepressants"],
            
            # Personality Disorders
            "Personality Disorders": ["borderline personality", "narcissistic personality", "antisocial personality", "avoidant personality", "dependent personality", "obsessive-compulsive personality", "paranoid personality", "schizoid personality", "schizotypal personality", "personality traits", "personality development"],
            
            # Eating Disorders
            "Eating Disorders": ["anorexia", "bulimia", "binge eating", "eating disorder", "body dysmorphia", "food restriction", "purging", "compulsive eating", "emotional eating", "disordered eating", "weight concerns", "body image", "nutrition", "meal planning", "recovery"],
            
            # Substance Use & Addiction
            "Substance Use & Addiction": ["addiction", "substance abuse", "alcoholism", "drug addiction", "recovery", "sobriety", "relapse", "withdrawal", "tolerance", "dependence", "12-step programs", "AA", "NA", "addiction treatment", "harm reduction", "substance use disorder"],
            
            # Grief & Loss
            "Grief & Loss": ["grief", "loss", "bereavement", "death", "mourning", "stages of grief", "complicated grief", "anticipatory grief", "disenfranchised grief", "grief counseling", "bereavement support", "loss of loved one", "pet loss", "job loss", "relationship loss"],
            
            # Life Transitions
            "Life Transitions": ["life changes", "transition", "moving", "job change", "career change", "retirement", "graduation", "marriage", "divorce", "parenthood", "empty nest", "midlife crisis", "quarter-life crisis", "identity crisis", "personal development", "life purpose"],
            
            # Work & Career Mental Health
            "Work & Career Mental Health": ["work stress", "job satisfaction", "work-life balance", "career burnout", "workplace bullying", "harassment", "discrimination", "job insecurity", "unemployment", "career development", "professional identity", "work relationships", "leadership stress", "imposter syndrome"],
            
            # Physical Health & Mental Health
            "Physical Health & Mental Health": ["chronic illness", "chronic pain", "disability", "mental health", "physical health", "mind-body connection", "psychosomatic", "health anxiety", "medical trauma", "illness anxiety", "somatic symptoms", "conversion disorder", "psychogenic", "medical conditions"],
            
            # Gender & Sexuality
            "Gender & Sexuality": ["gender identity", "sexual orientation", "LGBTQ+", "coming out", "gender dysphoria", "sexual identity", "gender expression", "transgender", "non-binary", "queer", "homophobia", "transphobia", "discrimination", "affirmation", "gender therapy"],
            
            # Aging & Mental Health
            "Aging & Mental Health": ["aging", "elderly", "senior mental health", "dementia", "Alzheimer's", "cognitive decline", "memory loss", "aging anxiety", "retirement", "elder care", "caregiver stress", "generational trauma", "wisdom", "life review", "legacy"],
            
            # Children & Adolescent Mental Health
            "Children & Adolescent Mental Health": ["child mental health", "adolescent", "teen", "youth", "childhood", "developmental disorders", "autism", "ASD", "developmental delays", "childhood trauma", "parenting", "family therapy", "play therapy", "school counseling", "youth services"],
            
            # Family & Parenting
            "Family & Parenting": ["family", "parenting", "parent-child relationship", "family dynamics", "family therapy", "parenting stress", "postpartum", "pregnancy", "infertility", "adoption", "foster care", "blended family", "single parenting", "co-parenting", "family conflict"],
            
            # Spirituality & Religion
            "Spirituality & Religion": ["spirituality", "religion", "faith", "spiritual crisis", "religious trauma", "spiritual growth", "meaning", "purpose", "existential", "philosophy", "values", "beliefs", "spiritual practices", "prayer", "meditation", "mindfulness"],
            
            # Technology & Mental Health
            "Technology & Mental Health": ["social media", "internet addiction", "screen time", "digital wellness", "cyberbullying", "online harassment", "digital detox", "technology stress", "FOMO", "comparison", "digital mental health", "teletherapy", "mental health apps"],
            
            # Financial Mental Health
            "Financial Mental Health": ["financial stress", "money anxiety", "debt", "financial insecurity", "poverty", "economic hardship", "financial trauma", "money management", "financial planning", "financial therapy", "economic anxiety", "financial abuse"],
            
            # Environmental & Climate Mental Health
            "Environmental & Climate Mental Health": ["climate anxiety", "eco-anxiety", "environmental stress", "climate grief", "environmental trauma", "natural disasters", "climate change", "environmental health", "sustainability stress", "ecological grief", "environmental justice"],
            
            # Cultural & Identity Mental Health
            "Cultural & Identity Mental Health": ["cultural identity", "racial trauma", "discrimination", "racism", "microaggressions", "cultural stress", "acculturation", "immigration", "refugee", "cultural competence", "ethnic identity", "cultural heritage", "intersectionality"],
            
            # Neurodiversity
            "Neurodiversity": ["autism", "ASD", "ADHD", "dyslexia", "dyscalculia", "dyspraxia", "sensory processing", "neurodivergent", "neurotypical", "executive function", "sensory sensitivity", "special interests", "stimming", "neurodiversity movement", "accommodations"],
            
            # Recovery & Healing
            "Recovery & Healing": ["recovery", "healing", "rehabilitation", "treatment", "therapy", "counseling", "psychotherapy", "mental health treatment", "recovery journey", "healing process", "wellness recovery", "mental health recovery", "trauma recovery", "addiction recovery"],
            
            # Prevention & Education
            "Prevention & Education": ["mental health education", "prevention", "early intervention", "mental health awareness", "stigma reduction", "mental health literacy", "psychoeducation", "mental health promotion", "wellness education", "mental health first aid", "crisis prevention"],
            
            # Professional Help & Treatment
            "Professional Help & Treatment": ["therapy", "counseling", "psychotherapy", "psychiatrist", "psychologist", "therapist", "mental health professional", "treatment plan", "medication", "psychiatric medication", "mental health services", "mental health care", "therapeutic relationship"],
            
            # Support Systems & Community
            "Support Systems & Community": ["support group", "peer support", "community", "social support", "family support", "friend support", "mental health community", "support network", "advocacy", "mental health advocacy", "peer counseling", "mutual aid", "community mental health"],
            
            # Coping Skills & Strategies
            "Coping Skills & Strategies": ["coping skills", "coping strategies", "coping mechanisms", "stress management", "anxiety management", "depression management", "emotional regulation", "distress tolerance", "problem solving", "decision making", "time management", "boundary setting", "assertiveness"],
            
            # Positive Psychology
            "Positive Psychology": ["positive psychology", "happiness", "well-being", "flourishing", "resilience", "optimism", "gratitude", "hope", "meaning", "purpose", "strengths", "virtues", "positive emotions", "life satisfaction", "psychological well-being"],
            
            # Alternative & Complementary Therapies
            "Alternative & Complementary Therapies": ["art therapy", "music therapy", "dance therapy", "drama therapy", "equine therapy", "animal therapy", "nature therapy", "forest bathing", "acupuncture", "yoga", "tai chi", "qi gong", "herbal medicine", "holistic health", "integrative medicine"]
        }
    
    async def process_file(self, file_path: str, filename: str, *, chunk_size: int = 200, overlap: int = 30, mode: str = "chars", custom_keywords: Optional[List[str]] = None) -> Dict[str, Any]:
        """Process uploaded file"""
        file_extension = Path(filename).suffix.lower()
        
        if file_extension not in self.supported_extensions:
            raise ValueError(f"Unsupported file format: {file_extension}")
        
        # Extract content
        text_content = await self._extract_text(file_path, file_extension)
        
        # Clean content
        cleaned_text = self._clean_text(text_content)
        
        # Classify content (merge user custom keywords)
        categories = self._classify_content(cleaned_text, custom_keywords=custom_keywords)
        
        # Split content
        chunks = self._split_text(cleaned_text, chunk_size=chunk_size, overlap=overlap, mode=mode)
        
        return {
            "filename": filename,
            "extension": file_extension,
            "original_text": text_content,
            "cleaned_text": cleaned_text,
            "categories": categories,
            "chunks": chunks,
            "chunk_count": len(chunks),
            "processed_at": datetime.now().isoformat()
        }
    
    async def _extract_text(self, file_path: str, extension: str) -> str:
        """Extract content based on file type"""
        if extension == '.txt' or extension == '.md':
            return await self._extract_text_file(file_path)
        elif extension == '.pdf':
            return await self._extract_pdf(file_path)
        elif extension == '.docx':
            return await self._extract_docx(file_path)
        elif extension == '.xlsx':
            return await self._extract_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file format: {extension}")
    
    async def _extract_text_file(self, file_path: str) -> str:
        """Extract plain text file content"""
        async with aiofiles.open(file_path, 'rb') as f:
            raw_data = await f.read()
            encoding = chardet.detect(raw_data)['encoding'] or 'utf-8'
        
        async with aiofiles.open(file_path, 'r', encoding=encoding) as f:
            return await f.read()
    
    async def _extract_pdf(self, file_path: str) -> str:
        """Extract PDF file content"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            raise ValueError(f"PDF file processing failed: {str(e)}")

    async def _extract_docx(self, file_path: str) -> str:
        """Extract Word document content"""
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            raise ValueError(f"Word document processing failed: {str(e)}")

    async def _extract_xlsx(self, file_path: str) -> str:
        """Extract Excel file content"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text
        except Exception as e:
            raise ValueError(f"Excel document processing failed: {str(e)}")

    def _clean_text(self, text: str) -> str:
        """Clean content"""
        # Remove excessive whitespace
        text = ' '.join(text.split())
        
        # Remove special characters
        import re
        text = re.sub(r'[^\u4e00-\u9fff\w\s.,!?;:()【】""''、。！？；：（）]', '', text)
        
        return text.strip()
    
    def _classify_content(self, text: str, custom_keywords: Optional[List[str]] = None) -> List[str]:
        """Classify content"""
        categories = []
        text_lower = text.lower()
        
        # Merge custom keywords into a temporary category (for classification)
        temp_categories = dict(self.mental_health_categories)
        if custom_keywords:
            temp_categories["User Custom"] = custom_keywords
        
        for category, keywords in temp_categories.items():
            for keyword in keywords:
                if keyword.lower() in text_lower:
                    categories.append(category)
                    break
        
        return list(set(categories)) if categories else ["General Mental Health"]
    
    def _split_text(self, text: str, chunk_size: int = 200, overlap: int = 30, mode: str = "chars") -> List[Dict[str, Any]]:
        """Split content into chunks"""
        # By character or word mode
        if mode == 'words':
            # Use jieba for Chinese word segmentation
            units = list(jieba.cut(text))
        else:
            units = list(text)
        
        chunks = []
        current_chunk: List[str] = []
        current_measure = 0  # Length measure: character count or word count
        
        for i, unit in enumerate(units):
            current_chunk.append(unit)
            # words mode measures by "word count"; chars mode measures by "character count"
            if mode == 'words':
                current_measure += 1
            else:
                current_measure += len(unit)
            
            if current_measure >= chunk_size:
                chunk_text = ''.join(current_chunk)
                chunks.append({
                    "id": len(chunks),
                    "text": chunk_text,
                    "length": len(chunk_text),
                    "word_count": len(current_chunk),
                    "start_index": i - len(current_chunk) + 1,
                    "end_index": i
                })
                
                # Overlap processing
                if overlap > 0:
                    overlap_units = current_chunk[-overlap:] if len(current_chunk) > overlap else current_chunk
                else:
                    overlap_units = []
                current_chunk = overlap_units
                if mode == 'words':
                    current_measure = len(overlap_units)
                else:
                    current_measure = sum(len(u) for u in overlap_units)

        # Process the last chunk
        if current_chunk:
            chunk_text = ''.join(current_chunk)
            chunks.append({
                "id": len(chunks),
                "text": chunk_text,
                "length": len(chunk_text),
                "word_count": len(current_chunk),
                "start_index": len(units) - len(current_chunk),
                "end_index": len(units) - 1
            })
        
        return chunks

class MentalHealthChromaDBService:
    """Mental Health ChromaDB Vector Database Service"""
    
    def __init__(self, persist_directory: str = "./mental_health_chroma_db"):
        self.persist_directory = persist_directory
        self.client = None
        self.collection = None
        self.embedder = None
        self._initialize()
    
    def _initialize(self):
        """Initialize ChromaDB and embedding model"""
        # Create ChromaDB Client
        os.makedirs(self.persist_directory, exist_ok=True)
        self.client = chromadb.PersistentClient(path=self.persist_directory)

        # Get or create collection - use cosine similarity
        self.collection = self.client.get_or_create_collection(
            name="mental_health_knowledge_base",
            metadata={"description": "Mental health knowledge base document vector storage", "hnsw:space": "cosine"}
        )

        # Initialize embedding model
        try:
            self.embedder = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
        except Exception:
            # Fallback option
            self.embedder = SentenceTransformer('all-MiniLM-L6-v2')
    
    async def add_document(self, doc_id: str, chunks: List[Dict[str, Any]], metadata: Dict[str, Any]) -> bool:
        """Add document to vector database"""
        try:
            # Prepare data
            documents = []
            ids = []
            metadatas = []
            
            for chunk in chunks:
                chunk_id = f"{doc_id}_{chunk['id']}"
                documents.append(chunk['text'])
                ids.append(chunk_id)
                # Serialize category list to string to comply with Chroma's primitive type requirements
                categories_value = metadata.get('categories', [])
                if isinstance(categories_value, list):
                    categories_csv = ",".join(categories_value)
                    primary_category = categories_value[0] if categories_value else "General Mental Health"
                else:
                    categories_csv = str(categories_value) if categories_value is not None else ""
                    primary_category = categories_csv or "General Mental Health"

                metadatas.append({
                    "doc_id": doc_id,
                    "chunk_id": chunk['id'],
                    "filename": metadata.get('filename', ''),
                    "file_type": metadata.get('extension', ''),
                    "categories_csv": categories_csv,
                    "category": primary_category,
                    "mode": metadata.get('mode', 'chars'),
                    "chunk_size": metadata.get('chunk_size', 200),
                    "overlap": metadata.get('overlap', 30),
                    "chunk_length": chunk['length'],
                    "word_count": chunk['word_count'],
                    "chunk_type": chunk.get('chunk_type', 'default'),
                    "chunking_strategy": metadata.get('chunking_strategy', 'fixed_length'),
                    "created_at": datetime.now().isoformat()
                })
            
            # Generate embedding vectors
            embeddings = self.embedder.encode(documents, normalize_embeddings=True).tolist()
            
            # Add to ChromaDB
            self.collection.add(
                documents=documents,
                embeddings=embeddings,
                metadatas=metadatas,
                ids=ids
            )
            
            return True
        except Exception as e:
            print(f"Failed to add document to vector database: {str(e)}")
            return False
    
    async def search_similar(self, query: str, top_k: int = 5, category_filter: Optional[str] = None) -> List[Dict[str, Any]]:
        """Search similar document chunks"""
        try:
            # Generate query vector
            query_embedding = self.embedder.encode([query], normalize_embeddings=True).tolist()[0]
            
            # Build query conditions
            where_clause = None
            if category_filter:
                # Use contains filter for comma-separated strings
                where_clause = {"categories_csv": {"$contains": category_filter}}
            
            # Execute search
            results = self.collection.query(
                query_embeddings=[query_embedding],
                n_results=top_k,
                where=where_clause,
                include=['documents', 'metadatas', 'distances']
            )

            # Format results
            search_results = []
            if results['documents'] and results['documents'][0]:
                for i in range(len(results['documents'][0])):
                    distance = results['distances'][0][i]
                    # Convert cosine distance to similarity
                    distance = max(0, min(2, distance))
                    similarity = 1 - (distance / 2)
                    
                    search_results.append({
                        "id": results['ids'][0][i],
                        "text": results['documents'][0][i],
                        "metadata": results['metadatas'][0][i],
                        "similarity": similarity,
                        "distance": distance
                    })
            
            return search_results
        except Exception as e:
            print(f"Search failed: {str(e)}")
            return []
    
    async def search_by_category(self, category: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """Search documents by category"""
        try:
            results = self.collection.get(
                where={"categories_csv": {"$contains": category}},
                include=['documents', 'metadatas', 'ids']
            )
            
            search_results = []
            if results['documents']:
                for i, doc in enumerate(results['documents']):
                    metadata = results['metadatas'][i]
                    search_results.append({
                        "id": results['ids'][i],
                        "text": doc,
                        "metadata": metadata,
                        "similarity": 1.0,  # Category match similarity set to 1.0
                        "distance": 0.0
                    })
            
            return search_results
        except Exception as e:
            print(f"Category search failed: {str(e)}")
            return []
    
    async def delete_document(self, doc_id: str) -> bool:
        """Delete document"""
        try:
            # Find all related chunks
            results = self.collection.get(
                where={"doc_id": doc_id},
                include=['metadatas']
            )
            
            if results['ids']:
                # Delete all related chunks
                self.collection.delete(ids=results['ids'])
                return True
            return False
        except Exception as e:
            print(f"Failed to delete document: {str(e)}")
            return False
    
    async def get_all_documents(self) -> List[Dict[str, Any]]:
        """Get all document information"""
        try:
            results = self.collection.get(include=['metadatas'])

            # Group by document ID
            docs = {}
            for i, metadata in enumerate(results['metadatas']):
                doc_id = metadata['doc_id']
                if doc_id not in docs:
                    docs[doc_id] = {
                        "doc_id": doc_id,
                        "filename": metadata['filename'],
                        "file_type": metadata['file_type'],
                        # Restore stored string to list for frontend display
                        "categories": (
                            [c for c in str(metadata.get('categories_csv', '')).split(',') if c]
                        ),
                        "chunks": [],
                        "total_chunks": 0,
                        "chunking_strategy": metadata.get('chunking_strategy', 'fixed_length'),
                        "created_at": metadata['created_at']
                    }
                
                docs[doc_id]['chunks'].append({
                    "chunk_id": metadata['chunk_id'],
                    "length": metadata['chunk_length'],
                    "word_count": metadata['word_count']
                })
                docs[doc_id]['total_chunks'] += 1
            
            return list(docs.values())
        except Exception as e:
            print(f"Failed to get document list: {str(e)}")
            return []
    
    async def get_document_chunks(self, doc_id: str) -> List[Dict[str, Any]]:
        """Get all chunks of a document"""
        try:
            results = self.collection.get(
                where={"doc_id": doc_id},
                include=['documents', 'metadatas']
            )
            
            chunks = []
            if results['documents']:
                for i, doc in enumerate(results['documents']):
                    metadata = results['metadatas'][i]
                    chunks.append({
                        "chunk_id": metadata['chunk_id'],
                        "text": doc,
                        "length": metadata['chunk_length'],
                        "word_count": metadata['word_count'],
                        "created_at": metadata['created_at'],
                        "mode": metadata.get('mode', 'chars'),
                        "chunk_size": metadata.get('chunk_size', 200),
                        "overlap": metadata.get('overlap', 30)
                    })
            
            # Sort by chunk_id
            chunks.sort(key=lambda x: x['chunk_id'])
            return chunks
        except Exception as e:
            print(f"Failed to get document chunks: {str(e)}")
            return []

class MentalHealthRAGService:
    """Mental Health RAG Integration Service"""
    
    def __init__(self):
        self.doc_processor = MentalHealthDocumentProcessor()
        self.vector_db = MentalHealthChromaDBService()
        self.upload_dir = "./mental_health_uploads"
        os.makedirs(self.upload_dir, exist_ok=True)
    
    async def upload_and_process_document(self, file_content: bytes, filename: str, *, chunk_size: int = 200, overlap: int = 30, mode: str = "chars", custom_keywords: Optional[List[str]] = None) -> Dict[str, Any]:
        """Upload and process document"""
        # Generate unique document ID
        doc_id = str(uuid.uuid4())
        
        # Save file
        file_path = os.path.join(self.upload_dir, f"{doc_id}_{filename}")
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(file_content)
        
        try:
            # Process document
            processed_doc = await self.doc_processor.process_file(
                file_path,
                filename,
                chunk_size=chunk_size,
                overlap=overlap,
                mode=mode,
                custom_keywords=custom_keywords
            )

            # Add to vector database
            metadata = {
                "filename": filename,
                "extension": processed_doc["extension"],
                "categories": processed_doc["categories"],
                "doc_id": doc_id,
                "mode": mode,
                "chunk_size": chunk_size,
                "overlap": overlap
            }
            
            success = await self.vector_db.add_document(
                doc_id=doc_id,
                chunks=processed_doc["chunks"],
                metadata=metadata
            )
            
            if success:
                # Delete temporary file
                os.remove(file_path)
                
                return {
                    "success": True,
                    "doc_id": doc_id,
                    "filename": filename,
                    "categories": processed_doc["categories"],
                    "chunk_count": processed_doc["chunk_count"],
                    "processed_at": processed_doc["processed_at"],
                    "message": "Mental health document uploaded and processed successfully",
                    "params": {
                        "chunk_size": chunk_size,
                        "overlap": overlap,
                        "mode": mode,
                        "custom_keywords": custom_keywords or []
                    }
                }
            else:
                # Cleanup on failure
                os.remove(file_path)
                return {
                    "success": False,
                    "message": "Vectorization processing failed"
                }
        except Exception as e:
            # Cleanup on failure
            if os.path.exists(file_path):
                os.remove(file_path)
            return {
                "success": False,
                "message": f"Document processing failed: {str(e)}"
            }
    
    async def search_knowledge_base(self, query: str, top_k: int = 5, category_filter: Optional[str] = None) -> List[Dict[str, Any]]:
        """Search mental health knowledge base"""
        return await self.vector_db.search_similar(query, top_k, category_filter)
    
    async def search_by_category(self, category: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """Search mental health documents by category"""
        return await self.vector_db.search_by_category(category, top_k)
    
    async def generate_mental_health_response(self, query: str, context_chunks: List[Dict[str, Any]]) -> str:
        """Generate mental health response based on retrieved context"""
        # Build context
        context = "\n\n".join([chunk["text"] for chunk in context_chunks])

        # Build prompt
        prompt = f"""Based on the following mental health related information, answer the user's question:

Context:
{context}

User Question: {query}

Please provide a professional, warm, and empathetic response based on the context information. The response should:
1. Acknowledge and understand the user's feelings
2. Provide practical advice and strategies
3. Encourage seeking professional help if needed
4. Convey hope and the possibility of positive change

Response:"""
        
        return prompt
    
    async def delete_document(self, doc_id: str) -> bool:
        """Delete document"""
        return await self.vector_db.delete_document(doc_id)
    
    async def get_all_documents(self) -> List[Dict[str, Any]]:
        """Get all documents"""
        return await self.vector_db.get_all_documents()
    
    async def get_document_chunks(self, doc_id: str) -> List[Dict[str, Any]]:
        """Get document chunks"""
        return await self.vector_db.get_document_chunks(doc_id)
    
    async def get_available_categories(self) -> List[str]:
        """Get available document categories"""
        return list(self.doc_processor.mental_health_categories.keys())

# Global mental health RAG service instance
mental_health_rag_service = MentalHealthRAGService()
