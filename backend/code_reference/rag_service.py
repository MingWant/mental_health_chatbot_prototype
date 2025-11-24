"""
RAG (Retrieval-Augmented Generation) 服务
用于文档處理、向量化、檢索和答案生成
"""

import os
import uuid
import asyncio
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import chromadb
from chromadb.config import Settings
import aiofiles
import chardet
import jieba
from datetime import datetime
import json

# 文檔處理相關
import PyPDF2
from docx import Document
import openpyxl
from sentence_transformers import SentenceTransformer


class DocumentProcessor:
    """文檔處理器"""
    
    def __init__(self):
        self.supported_extensions = {'.txt', '.pdf', '.docx', '.xlsx', '.md'}
    
    async def process_file(self, file_path: str, filename: str) -> Dict[str, Any]:
        """處理Upload的文件"""
        file_extension = Path(filename).suffix.lower()
        
        if file_extension not in self.supported_extensions:
            raise ValueError(f"不支持File的格式: {file_extension}")
        
        # 提取內容
        text_content = await self._extract_text(file_path, file_extension)
        
        # 清洗內容
        cleaned_text = self._clean_text(text_content)

        # 拆分內容
        chunks = self._split_text(cleaned_text)
        
        return {
            "filename": filename,
            "extension": file_extension,
            "original_text": text_content,
            "cleaned_text": cleaned_text,
            "chunks": chunks,
            "chunk_count": len(chunks),
            "processed_at": datetime.now().isoformat()
        }
    
    async def _extract_text(self, file_path: str, extension: str) -> str:
        """根據文件類型提取內容"""
        if extension == '.txt' or extension == '.md':
            return await self._extract_text_file(file_path)
        elif extension == '.pdf':
            return await self._extract_pdf(file_path)
        elif extension == '.docx':
            return await self._extract_docx(file_path)
        elif extension == '.xlsx':
            return await self._extract_xlsx(file_path)
        else:
            raise ValueError(f"不支持的File格式: {extension}")
    
    async def _extract_text_file(self, file_path: str) -> str:
        """提取純文本文件內容"""
        async with aiofiles.open(file_path, 'rb') as f:
            raw_data = await f.read()
            # 檢測編碼
            encoding = chardet.detect(raw_data)['encoding'] or 'utf-8'
        
        async with aiofiles.open(file_path, 'r', encoding=encoding) as f:
            return await f.read()
    
    async def _extract_pdf(self, file_path: str) -> str:
        """提取PDF文件內容"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            raise ValueError(f"PDF文件處理失敗: {str(e)}")

    async def _extract_docx(self, file_path: str) -> str:
        """提取Word文檔內容"""
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            raise ValueError(f"Word文檔處理失敗: {str(e)}")

    async def _extract_xlsx(self, file_path: str) -> str:
        """提取Excel文件內容"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Table: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text
        except Exception as e:
            raise ValueError(f"Excel文檔處理失敗: {str(e)}")

    def _clean_text(self, text: str) -> str:
        """清洗內容"""
        # 去除多余的空白符號
        text = ' '.join(text.split())
        
        # 去除特殊符號
        import re
        text = re.sub(r'[^\u4e00-\u9fff\w\s.,!?;:()【】""''、。！？；：（）]', '', text)
        
        return text.strip()
    
    def _split_text(self, text: str, chunk_size: int = 200, overlap: int = 30) -> List[Dict[str, Any]]:
        """将内容拆分为Chunk"""
        # 使用jieba進行中文分詞
        words = list(jieba.cut(text))
        
        chunks = []
        current_chunk = []
        current_length = 0
        
        for i, word in enumerate(words):
            current_chunk.append(word)
            current_length += len(word)
            
            if current_length >= chunk_size:
                chunk_text = ''.join(current_chunk)
                chunks.append({
                    "id": len(chunks),
                    "text": chunk_text,
                    "length": len(chunk_text),
                    "word_count": len(current_chunk),
                    "start_index": i - len(current_chunk) + 1,
                    "end_index": i
                })
                
                # 重疊處理（overlap）
                overlap_words = current_chunk[-overlap:] if len(current_chunk) > overlap else current_chunk
                current_chunk = overlap_words
                current_length = sum(len(word) for word in overlap_words)

        # 處理最後一個Chunk
        if current_chunk:
            chunk_text = ''.join(current_chunk)
            chunks.append({
                "id": len(chunks),
                "text": chunk_text,
                "length": len(chunk_text),
                "word_count": len(current_chunk),
                "start_index": len(words) - len(current_chunk),
                "end_index": len(words) - 1
            })
        
        return chunks


class ChromaDBService:
    """ChromaDB 向量數據庫服務(Vector Database)"""
    
    def __init__(self, persist_directory: str = "./chroma_db"):
        self.persist_directory = persist_directory
        self.client = None
        self.collection = None
        self.embedder = None
        self._initialize()
    
    def _initialize(self):
        """Initialize ChromaDB和嵌入模型"""
        # 創建ChromaDB的Client
        os.makedirs(self.persist_directory, exist_ok=True)
        self.client = chromadb.PersistentClient(path=self.persist_directory)

        # 獲取或創建集合 - 使用余弦相似度
        self.collection = self.client.get_or_create_collection(
            name="knowledge_base",
            metadata={"description": "知識庫文檔向量存儲", "hnsw:space": "cosine"}
        )

        # Initialize嵌入模型
        try:
            self.embedder = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
        except Exception:
            # 備選方案
            self.embedder = SentenceTransformer('all-MiniLM-L6-v2')
    
    async def add_document(self, doc_id: str, chunks: List[Dict[str, Any]], metadata: Dict[str, Any]) -> bool:
        """添加文檔到向量數據庫"""
        try:
            # 準備數據
            documents = []
            ids = []
            metadatas = []
            
            for chunk in chunks:
                chunk_id = f"{doc_id}_{chunk['id']}"
                documents.append(chunk['text'])
                ids.append(chunk_id)
                metadatas.append({
                    "doc_id": doc_id,
                    "chunk_id": chunk['id'],
                    "filename": metadata.get('filename', ''),
                    "file_type": metadata.get('extension', ''),
                    "chunk_length": chunk['length'],
                    "word_count": chunk['word_count'],
                    "created_at": datetime.now().isoformat()
                })
            
            # 生成嵌入向量
            embeddings = self.embedder.encode(documents, normalize_embeddings=True).tolist()
            
            # 添加到ChromaDB
            self.collection.add(
                documents=documents,
                embeddings=embeddings,
                metadatas=metadatas,
                ids=ids
            )
            
            return True
        except Exception as e:
            print(f"添加文檔到向量數據庫失敗: {str(e)}")
            return False
    
    async def search_similar(self, query: str, top_k: int = 5) -> List[Dict[str, Any]]:
        """搜索相似文檔塊"""
        try:
            # 生成查詢向量
            query_embedding = self.embedder.encode([query], normalize_embeddings=True).tolist()[0]
            
            # 執行搜索
            results = self.collection.query(
                query_embeddings=[query_embedding],
                n_results=top_k,
                include=['documents', 'metadatas', 'distances']
            )

            # 格式化結果
            search_results = []
            if results['documents'] and results['documents'][0]:
                for i in range(len(results['documents'][0])):
                    distance = results['distances'][0][i]
                    # 余弦距離轉換為相似度：cosine_similarity = 1 - cosine_distance
                    # 但ChromaDB的余弦距離範圍是[0,2]，需要限制在合理範圍
                    distance = max(0, min(2, distance))  # 確保距離在[0,2]範圍
                    similarity = 1 - (distance / 2)  # 轉換為[0,1]範圍的相似度
                    
                    search_results.append({
                        "id": results['ids'][0][i],
                        "text": results['documents'][0][i],
                        "metadata": results['metadatas'][0][i],
                        "similarity": similarity,
                        "distance": distance
                    })
            
            return search_results
        except Exception as e:
            print(f"搜索失敗: {str(e)}")
            return []
    
    async def delete_document(self, doc_id: str) -> bool:
        """刪除文檔"""
        try:
            # 查找所有相關的chunk
            results = self.collection.get(
                where={"doc_id": doc_id},
                include=['metadatas']
            )
            
            if results['ids']:
                # 刪除所有相關Chunk
                self.collection.delete(ids=results['ids'])
                return True
            return False
        except Exception as e:
            print(f"刪除文檔失敗: {str(e)}")
            return False
    
    async def get_all_documents(self) -> List[Dict[str, Any]]:
        """獲取所有文檔信息"""
        try:
            results = self.collection.get(include=['metadatas'])

            # 按文檔ID分組
            docs = {}
            for i, metadata in enumerate(results['metadatas']):
                doc_id = metadata['doc_id']
                if doc_id not in docs:
                    docs[doc_id] = {
                        "doc_id": doc_id,
                        "filename": metadata['filename'],
                        "file_type": metadata['file_type'],
                        "chunks": [],
                        "total_chunks": 0,
                        "created_at": metadata['created_at']
                    }
                
                docs[doc_id]['chunks'].append({
                    "chunk_id": metadata['chunk_id'],
                    "length": metadata['chunk_length'],
                    "word_count": metadata['word_count']
                })
                docs[doc_id]['total_chunks'] += 1
            
            return list(docs.values())
        except Exception as e:
            print(f"獲取文檔列表失敗: {str(e)}")
            return []
    
    async def get_document_chunks(self, doc_id: str) -> List[Dict[str, Any]]:
        """獲取文檔的所有Chunk"""
        try:
            results = self.collection.get(
                where={"doc_id": doc_id},
                include=['documents', 'metadatas']
            )
            
            chunks = []
            if results['documents']:
                for i, doc in enumerate(results['documents']):
                    metadata = results['metadatas'][i]
                    chunks.append({
                        "chunk_id": metadata['chunk_id'],
                        "text": doc,
                        "length": metadata['chunk_length'],
                        "word_count": metadata['word_count'],
                        "created_at": metadata['created_at']
                    })
            
            # 按chunk_id排序
            chunks.sort(key=lambda x: x['chunk_id'])
            return chunks
        except Exception as e:
            print(f"獲取文檔Chunk失敗: {str(e)}")
            return []


class RAGService:
    """RAG集成服務"""
    
    def __init__(self):
        self.doc_processor = DocumentProcessor()
        self.vector_db = ChromaDBService()
        self.upload_dir = "./uploads"
        os.makedirs(self.upload_dir, exist_ok=True)
    
    async def upload_and_process_document(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """上傳並處理文檔"""
        # 生成唯一文檔ID
        doc_id = str(uuid.uuid4())
        
        # 保存文件
        file_path = os.path.join(self.upload_dir, f"{doc_id}_{filename}")
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(file_content)
        
        try:
            # 處理文檔
            processed_doc = await self.doc_processor.process_file(file_path, filename)

            # 添加到向量數據庫
            metadata = {
                "filename": filename,
                "extension": processed_doc["extension"],
                "doc_id": doc_id
            }
            
            success = await self.vector_db.add_document(
                doc_id=doc_id,
                chunks=processed_doc["chunks"],
                metadata=metadata
            )
            
            if success:
                # 刪除臨時文件
                os.remove(file_path)
                
                return {
                    "success": True,
                    "doc_id": doc_id,
                    "filename": filename,
                    "chunk_count": processed_doc["chunk_count"],
                    "processed_at": processed_doc["processed_at"],
                    "message": "文檔上傳並處理成功"
                }
            else:
                # 清理失敗
                os.remove(file_path)
                return {
                    "success": False,
                    "message": "向量化處理失敗"
                }
        except Exception as e:
            # 清理失敗
            if os.path.exists(file_path):
                os.remove(file_path)
            return {
                "success": False,
                "message": f"文檔處理失敗: {str(e)}"
            }
    
    async def search_knowledge_base(self, query: str, top_k: int = 5) -> List[Dict[str, Any]]:
        """搜索知識庫"""
        return await self.vector_db.search_similar(query, top_k)
    
    async def generate_rag_response(self, query: str, context_chunks: List[Dict[str, Any]]) -> str:
        """基于檢索到的上下文生成回答"""
        # 構建上下文
        context = "\n\n".join([chunk["text"] for chunk in context_chunks])

        # 構建prompt
        prompt = f"""基于以下上下文信息回答用戶問題：

上下文：
{context}

用戶問題：{query}

請基於上下文信息回答問題。如果上下文中沒有相關信息，請說明無法找到相關信息。

回答："""
        
        return prompt
    
    async def delete_document(self, doc_id: str) -> bool:
        """刪除文檔"""
        return await self.vector_db.delete_document(doc_id)
    
    async def get_all_documents(self) -> List[Dict[str, Any]]:
        """獲取所有文檔"""
        return await self.vector_db.get_all_documents()
    
    async def get_document_chunks(self, doc_id: str) -> List[Dict[str, Any]]:
        """獲取文檔的拆分塊"""
        return await self.vector_db.get_document_chunks(doc_id)


# 全局RAG服務實例
rag_service = RAGService()
